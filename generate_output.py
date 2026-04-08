import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json, datetime, calendar, re

WBS_MAP = {
    "APEL":        "APEL-SSA-TAR-ATA-HIRATE-MC-199",
    "BFHL":        "BFHL-SSA-TAR-ATA-HIRATE-MC-210",
    "BWHPL":       "BWHPL-SSA-TAR-ATA-HIRATE-MC-340",
    "DATL":        "DATL-SSA-TAR-ATA-HIRATE-MC-198",
    "FRHL_HIRATE": "FRHL-SSA-TAR-ATA-HIRATE-MC-197",
    "FRHL_ENEXCO": "FRHL-SSA-TAR-ASA-E. NSV ANALYSIS-MC-147",
    "GAEPL":       "GAPEL-SSA-TAR-ATA-HIRATE-MC-200",
    "JMTPL":       "JMTL-SSA-TAR-ATA-HIRATE-MC-194",
    "KETPL":       "KETL-SSA-TAR-ATA-HIRATE-MC-196",
    "KMTPL":       "KMTL-SSA-TAR-ATA-HIRATE-MC-195",
    "KTIPL":       "KTIPL-SSA-TAR-ATA-HIRATE-MC-348",
    "MBEL":        "MBEL-SSA-TAR-ATA-HIRATE-MC-192",
    "MHPL":        "MHPL-SSA-TAR-ATA-HIRATE-MC-346",
    "MKTL":        "MKTL-SSA-TAR-ATA-HIRATE-MC-191",
    "MSHP":        "MSHPL-SSA-TAR-ATA-HIRATE-MC-343",
    "NAM":         "NAMEL-SSA-TAR-ATA-HIRATE-MC-400",
    "NDEPL":       "NDEPL-SSA-TAR-ATA-HIRATE-MC-190",
    "NKTPL":       "NKTL-SSA-TAR-ATA-HIRATE-MC-189",
    "SIPL":        "SIPL-SSA-TAR-ATA-HIRATE-MC-350",
    "SMTPL":       "SMTL-SSA-TAR-ATA-HIRATE-MC-188",
    "SPPL":        "SPPL-SSA-TAR-ATA-HIRATE-MC-352",
    "WVEL":        "WVEPL-SSA-TAR-ATA-HIRATE-MC-193",
    "WUPTL":       "WUPTL-SSA-TAR-ATA-HIRATE-MC-187",
    "Lightstorm":  "LIGHTSTORM-SSA-ATMS-INST-LGS-01-MC-472",
}

WBS_COLUMN_ORDER = [
    "APEL-SSA-TAR-ATA-HIRATE-MC-199",
    "BFHL-SSA-TAR-ATA-HIRATE-MC-210",
    "BWHPL-SSA-TAR-ATA-HIRATE-MC-340",
    "DATL-SSA-TAR-ATA-HIRATE-MC-198",
    "FRHL-SSA-TAR-ATA-HIRATE-MC-197",
    "GAPEL-SSA-TAR-ATA-HIRATE-MC-200",
    "JMTL-SSA-TAR-ATA-HIRATE-MC-194",
    "KETL-SSA-TAR-ATA-HIRATE-MC-196",
    "KMTL-SSA-TAR-ATA-HIRATE-MC-195",
    "KTIPL-SSA-TAR-ATA-HIRATE-MC-348",
    "MBEL-SSA-TAR-ATA-HIRATE-MC-192",
    "MHPL-SSA-TAR-ATA-HIRATE-MC-346",
    "MKTL-SSA-TAR-ATA-HIRATE-MC-191",
    "MSHPL-SSA-TAR-ATA-HIRATE-MC-343",
    "NAMEL-SSA-TAR-ATA-HIRATE-MC-400",
    "NDEPL-SSA-TAR-ATA-HIRATE-MC-190",
    "NKTL-SSA-TAR-ATA-HIRATE-MC-189",
    "SIPL-SSA-TAR-ATA-HIRATE-MC-350",
    "SMTL-SSA-TAR-ATA-HIRATE-MC-188",
    "SPPL-SSA-TAR-ATA-HIRATE-MC-352",
    "WVEPL-SSA-TAR-ATA-HIRATE-MC-193",
    "WUPTL-SSA-TAR-ATA-HIRATE-MC-187",
    "FRHL-SSA-TAR-ASA-E. NSV ANALYSIS-MC-147",
    "LIGHTSTORM-SSA-ATMS-INST-LGS-01-MC-472",
]

def norm(s):
    return re.sub(r'\s+', ' ', str(s).strip().lower())

def _s(row, col):
    v = row.get(col, '')
    return '' if pd.isna(v) or str(v).strip() in ('nan','NaN','') else str(v).strip()

def parse_input():
    raw = pd.read_excel('emp_data_input.xlsx', sheet_name='Sheet1', header=None)
    hirate_projects, enexco_projects, lightstorm_projects = set(), set(), set()
    for i in range(2, 23):
        code = raw.iloc[i, 0]
        if pd.isna(code) or not str(code).strip(): continue
        code = str(code).strip()
        if pd.notna(raw.iloc[i, 1]): hirate_projects.add(code)
        if pd.notna(raw.iloc[i, 2]): enexco_projects.add(code)
        if pd.notna(raw.iloc[i, 4]): lightstorm_projects.add(code)

    employees = []
    for i in range(2, len(raw)):
        name = raw.iloc[i, 9]
        if pd.isna(name) or not str(name).strip(): continue
        name = str(name).strip()
        attendance = raw.iloc[i, 10]
        if pd.isna(attendance): continue
        attendance = int(attendance)
        projects = []
        for p_start in [11, 13, 15, 17]:
            proj_code = raw.iloc[i, p_start]
            proj_days = raw.iloc[i, p_start + 1]
            if pd.notna(proj_code) and pd.notna(proj_days):
                code = str(proj_code).strip()
                days = int(proj_days)
                if code in lightstorm_projects or code == 'Lightstorm':
                    wbs = WBS_MAP['Lightstorm']
                elif code == 'FRHL' and code in enexco_projects:
                    wbs = WBS_MAP['FRHL_ENEXCO']
                elif code == 'FRHL':
                    wbs = WBS_MAP['FRHL_HIRATE']
                elif code in WBS_MAP:
                    wbs = WBS_MAP[code]
                else:
                    wbs = code
                projects.append({'code': code, 'days': days, 'wbs': wbs})
        employees.append({'name': name, 'attendance': attendance, 'projects': projects})
    return employees

def parse_sheet2():
    df = pd.read_excel('emp_data_input.xlsx', sheet_name='Sheet2', dtype=str)
    records = {}
    for _, row in df.iterrows():
        name = str(row['Service Provider']).strip()
        sig_path = _s(row, 'Path')
        sig_filename = sig_path.replace('\\', '/').split('/')[-1] if sig_path else ''
        rec = {
            'service_provider': name,
            'address':        _s(row, 'Address'),
            'email':          _s(row, 'Email'),
            'contact':        _s(row, 'Contact No'),
            'pan':            _s(row, 'PAN No'),
            'account_name':   _s(row, 'Account-Name'),
            'bank_name':      _s(row, 'Bank Name'),
            'account_number': _s(row, 'Bank Account Number'),
            'ifsc':           _s(row, 'IFSC Code'),
            'sig_filename':   sig_filename,
        }
        records[norm(name)] = rec
    return records

# Manual alias map for name mismatches between Sheet1 and Sheet2
NAME_ALIASES = {
    "gaja balanarayana":        "gaja bala narayana",
    "m.esthar rani":            "meka esthar rani",
    "neeli srivani":            "neeli sreevani",
    "thabasum afreen":          "tabassum afreen",
    "thurpati vijay bhaskar":   "thurpati vijay baskar",
}

def find_provider(name, sheet2):
    key = norm(name)
    key = NAME_ALIASES.get(key, key)
    if key in sheet2: return sheet2[key]
    key2 = re.sub(r'[.\s]+', ' ', key).strip()
    for k, v in sheet2.items():
        if re.sub(r'[.\s]+', ' ', k).strip() == key2: return v
    key_tokens = set(key2.split())
    for k, v in sheet2.items():
        k_tokens = set(re.sub(r'[.\s]+', ' ', k).strip().split())
        if key_tokens == k_tokens or key_tokens.issubset(k_tokens) or k_tokens.issubset(key_tokens):
            return v
    return {}

BLUE_FILL   = PatternFill("solid", fgColor="BDD7EE")
HEADER_FILL = PatternFill("solid", fgColor="2E75B6")
GREY_FILL   = PatternFill("solid", fgColor="D9D9D9")
ALT_FILL    = PatternFill("solid", fgColor="EBF3FB")
WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")
ORANGE_FILL = PatternFill("solid", fgColor="FFC000")
RED_FILL    = PatternFill("solid", fgColor="FFE6E6")

def thin_border():
    s = Side(style='thin')
    return Border(left=s, right=s, top=s, bottom=s)

def set_cell(ws, row, col, value, bold=False, fill=None, align='center',
             border=True, wrap=False, sz=9, color="000000", fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name='Arial', bold=bold, size=sz, color=color)
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    if fill: c.fill = fill
    if border: c.border = thin_border()
    if fmt: c.number_format = fmt
    return c

def build_sheet1(ws, employees, sheet2):
    headers = [
        'S No.','Name','Attendance','Amount','Rate/Day','Days',
        'Project 1','Days 1','Project 2','Days 2','Project 3','Days 3',
        'No. of working days 1','WBS Elements 1','Payable Amount 1',
        'No. of working days 2','WBS Elements 2','Payable Amount 2',
        'No. of working days 3','WBS Elements 3','Payable Amount 3',
        'Total Working Days','Total Amount',
        'Service Provider','Address','Email','Contact No','PAN No',
        'Account-Name','Bank Name','Bank Account Number','IFSC Code'
    ]
    ws.row_dimensions[1].height = 30
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = Font(name='Arial', bold=True, size=9, color="FFFFFF")
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = thin_border()
    col_widths = [6,22,10,10,12,7,10,7,10,7,10,7,18,42,16,18,42,16,18,42,16,16,12,22,42,30,14,12,22,20,22,12]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    MONTHLY, MONTH_DAYS = 16500, 28
    for ri, emp in enumerate(employees, 2):
        att = emp['attendance']
        projs = emp['projects']
        rate = MONTHLY / MONTH_DAYS
        total = round(att * rate)
        prov = find_provider(emp['name'], sheet2)
        p1 = projs[0] if len(projs) > 0 else {}
        p2 = projs[1] if len(projs) > 1 else {}
        p3 = projs[2] if len(projs) > 2 else {}
        a1 = round(p1['days'] * rate) if p1 else ''
        a2 = round(p2['days'] * rate) if p2 else ''
        a3 = round(p3['days'] * rate) if p3 else ''
        row_data = [
            ri-1, emp['name'], att, total, round(rate,6), att,
            p1.get('code',''), p1.get('days','') if p1 else '',
            p2.get('code','') if p2 else '', p2.get('days','') if p2 else '',
            p3.get('code','') if p3 else '', p3.get('days','') if p3 else '',
            p1.get('days','') if p1 else '', p1.get('wbs','') if p1 else '', a1,
            p2.get('days','') if p2 else '', p2.get('wbs','') if p2 else '', a2,
            p3.get('days','') if p3 else '', p3.get('wbs','') if p3 else '', a3,
            att, total,
            prov.get('service_provider', emp['name']),
            prov.get('address',''), prov.get('email',''), prov.get('contact',''),
            prov.get('pan',''), prov.get('account_name',''), prov.get('bank_name',''),
            prov.get('account_number',''), prov.get('ifsc',''),
        ]
        fill = ALT_FILL if ri % 2 == 0 else WHITE_FILL
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = Font(name='Arial', size=9)
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=(ci in [2,14,17,20,24,25]))
            c.fill = fill
            c.border = thin_border()
            if ci in [27, 31]:  # Contact No, Bank Account: force text
                c.number_format = '@'
        ws.row_dimensions[ri].height = 20

def build_timesheet(ws, employees, sheet2):
    MONTHLY, MONTH_DAYS = 16500, 28
    rate = MONTHLY / MONTH_DAYS
    year, month = 2026, 2
    _, days_in_month = calendar.monthrange(year, month)
    dates = [datetime.date(year, month, d) for d in range(1, days_in_month+1)]
    day_abbr = ['Sun','Mon','Tue','Wed','Thu','Fri','Sat']

    date_start = 3
    date_end   = date_start + days_in_month - 1
    rate_col   = date_end + 1
    att_col    = rate_col + 1
    wbs_start  = att_col + 1
    pay_col    = wbs_start + len(WBS_COLUMN_ORDER)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=pay_col)
    c = ws.cell(row=1, column=1, value="TIME SHEET FEBRUARY 2026")
    c.font = Font(name='Arial', bold=True, size=14, color="FFFFFF")
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    ws.row_dimensions[2].height = 28
    for col, val in [(1,'S.No'),(2,'NAME')]:
        set_cell(ws, 2, col, val, bold=True, fill=HEADER_FILL, sz=9, color="FFFFFF")
    for di, d in enumerate(dates):
        col = date_start + di
        c = ws.cell(row=2, column=col, value=d)
        c.number_format = 'DD-MMM'
        c.font = Font(name='Arial', bold=True, size=8, color="FFFFFF")
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = thin_border()
    set_cell(ws, 2, rate_col, 'Rate',       bold=True, fill=HEADER_FILL, sz=9, color="FFFFFF")
    set_cell(ws, 2, att_col,  'Attendance', bold=True, fill=HEADER_FILL, sz=9, color="FFFFFF")
    for wi, wbs in enumerate(WBS_COLUMN_ORDER):
        col = wbs_start + wi
        c = ws.cell(row=2, column=col, value=wbs)
        c.font = Font(name='Arial', bold=True, size=8, color="FFFFFF")
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = thin_border()
        ws.column_dimensions[get_column_letter(col)].width = 18
    set_cell(ws, 2, pay_col, 'PAY', bold=True, fill=HEADER_FILL, sz=9, color="FFFFFF")

    ws.row_dimensions[3].height = 14
    set_cell(ws, 3, 1, '', fill=GREY_FILL)
    set_cell(ws, 3, 2, '', fill=GREY_FILL)
    for di, d in enumerate(dates):
        col = date_start + di
        display_dow = (d.weekday() + 1) % 7
        c = ws.cell(row=3, column=col, value=day_abbr[display_dow])
        c.font = Font(name='Arial', size=8)
        c.alignment = Alignment(horizontal='center')
        c.fill = GREY_FILL
        c.border = thin_border()
    for col in [rate_col, att_col] + list(range(wbs_start, pay_col+1)):
        set_cell(ws, 3, col, '', fill=GREY_FILL)

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 26
    for di in range(days_in_month):
        ws.column_dimensions[get_column_letter(date_start + di)].width = 4.5
    ws.column_dimensions[get_column_letter(rate_col)].width = 11
    ws.column_dimensions[get_column_letter(att_col)].width = 10
    ws.column_dimensions[get_column_letter(pay_col)].width = 13

    wbs_totals = {wbs: 0.0 for wbs in WBS_COLUMN_ORDER}
    grand_total = 0.0

    for ri, emp in enumerate(employees, 4):
        att = emp['attendance']
        projs = emp['projects']
        total = round(att * rate)
        fill = WHITE_FILL if ri % 2 == 0 else ALT_FILL

        set_cell(ws, ri, 1, ri-3, fill=fill, sz=9)
        set_cell(ws, ri, 2, emp['name'], fill=fill, sz=9, align='left')

        day_marks = []
        for d in dates:
            day_marks.append('H' if d.weekday() == 6 else None)
        work_slots = [i for i, m in enumerate(day_marks) if m is None]
        for idx in work_slots[:att]:  day_marks[idx] = 'P'
        for idx in work_slots[att:]:  day_marks[idx] = 'L'

        for di, mark in enumerate(day_marks):
            col = date_start + di
            c = ws.cell(row=ri, column=col, value=mark)
            c.font = Font(name='Arial', size=8)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = thin_border()
            c.fill = ORANGE_FILL if mark == 'H' else (RED_FILL if mark == 'L' else fill)

        set_cell(ws, ri, rate_col, round(rate,6), fill=fill, sz=9)
        set_cell(ws, ri, att_col, att,            fill=fill, sz=9)

        wbs_row = {wbs: 0.0 for wbs in WBS_COLUMN_ORDER}
        for p in projs:
            if p['wbs'] in wbs_row:
                wbs_row[p['wbs']] += round(p['days'] * rate, 6)
        for wi, wbs in enumerate(WBS_COLUMN_ORDER):
            col = wbs_start + wi
            val = round(wbs_row[wbs], 6)
            c = ws.cell(row=ri, column=col, value=val if val else 0.0)
            c.font = Font(name='Arial', size=9)
            c.alignment = Alignment(horizontal='right', vertical='center')
            c.border = thin_border()
            c.fill = fill
            if val > 0: c.number_format = '#,##0.00'
            wbs_totals[wbs] += val

        set_cell(ws, ri, pay_col, total, fill=fill, sz=9, align='right', fmt='#,##0')
        grand_total += total
        ws.row_dimensions[ri].height = 16

    tr = len(employees) + 4
    ws.row_dimensions[tr].height = 18
    set_cell(ws, tr, 1, '', bold=True, fill=BLUE_FILL)
    set_cell(ws, tr, 2, 'TOTAL', bold=True, fill=BLUE_FILL)
    for di in range(days_in_month):
        set_cell(ws, tr, date_start+di, '', fill=BLUE_FILL)
    set_cell(ws, tr, rate_col, '', fill=BLUE_FILL)
    set_cell(ws, tr, att_col, '', fill=BLUE_FILL)
    for wi, wbs in enumerate(WBS_COLUMN_ORDER):
        col = wbs_start + wi
        c = ws.cell(row=tr, column=col, value=round(wbs_totals[wbs], 2))
        c.font = Font(name='Arial', bold=True, size=9)
        c.alignment = Alignment(horizontal='right', vertical='center')
        c.border = thin_border()
        c.fill = BLUE_FILL
        c.number_format = '#,##0.00'
    c = ws.cell(row=tr, column=pay_col, value=grand_total)
    c.font = Font(name='Arial', bold=True, size=10)
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.border = thin_border()
    c.fill = BLUE_FILL
    c.number_format = '#,##0'
    ws.freeze_panes = 'C4'

def main():
    employees = parse_input()
    sheet2    = parse_sheet2()

    wb  = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = 'Sheet1'
    build_sheet1(ws1, employees, sheet2)
    ws_ts = wb.create_sheet('Time Sheet')
    build_timesheet(ws_ts, employees, sheet2)
    wb._sheets = [ws_ts, ws1]
    wb.save('Salary_TimeSheet_Output_new.xlsx')
    print("Saved Excel")

    MONTHLY, MONTH_DAYS = 16500, 28
    rate = MONTHLY / MONTH_DAYS
    emp_list = []
    for emp in employees:
        prov  = find_provider(emp['name'], sheet2)
        total = round(emp['attendance'] * rate)
        emp_list.append({**emp, **prov, 'total_amount': total, 'rate': round(rate,6)})

    with open('emp_data.json', 'w', encoding='utf-8') as f:
        json.dump(emp_list, f, ensure_ascii=False, indent=2)
    print("JSON exported")

    print("\n=== Provider match check ===")
    for e in emp_list:
        status = 'OK  ' if e.get('address') else 'MISS'
        print(f"{status} [{e['name']}] -> [{e.get('service_provider','')}]")

main()
